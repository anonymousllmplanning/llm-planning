#!/usr/bin/env python3
"""
Dual Facet Evaluation Runner  - Dataset-Aware Version

Key Improvements:
1. Uses ast_evaluation_system with dataset-aware metrics
2. Properly handles None values for N/A metrics (Delta has no arguments)
3. Computes separate aggregate metrics for datasets with/without arguments
4. Cleaner CSV output with N/A indicators

Usage:
    python dual_facet_evaluation_runner_v2.py \
        --input unified_50.qwen32b.jsonl \
        --output_csv per_sample.csv \
        --output_summary summary.json
"""

from __future__ import annotations
import argparse
import csv
import json
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional
from collections import Counter, defaultdict
from dataclasses import dataclass
import copy

# Use V2 evaluation system
from .metrics import ASTEvaluationSystem, _extract_tool_name
from src.inference.prompts import normalize_tool_environment

# Import utilities
from .utils import load_jsonl, mean, std, format_metric, print_summary, write_csv


def _load_gold_records(path: Path) -> Dict[str, Dict[str, Any]]:
    """Load canonical dataset records keyed by meta.id from JSON or JSONL."""
    if path.suffix.lower() == ".jsonl":
        records = list(load_jsonl(path))
    else:
        with path.open() as f:
            data = json.load(f)
        records = data if isinstance(data, list) else []

    by_id: Dict[str, Dict[str, Any]] = {}
    for rec in records:
        meta = rec.get("meta", {}) or {}
        sample_id = str(meta.get("id", ""))
        if sample_id:
            by_id[sample_id] = rec
    return by_id


def _record_id(record: Dict[str, Any]) -> str:
    meta = record.get("meta") or record.get("metadata") or {}
    return str(
        meta.get("id")
        or meta.get("sample_id")
        or record.get("id")
        or record.get("sample_id")
        or ""
    )


def _infer_gaia_async_path(gold_dataset: Optional[Path]) -> Optional[Path]:
    """Infer the async GAIA reference file matching a canonical GAIA dataset."""
    if not gold_dataset:
        return None
    text = str(gold_dataset)
    repo_root = Path(__file__).resolve().parents[2]
    for cat in ("A", "B", "C", "D"):
        if f"cat_{cat}" not in text and f"cat.{cat}" not in text:
            continue
        candidates = [
            repo_root / "Asynchronous" / f"gaia_cat_{cat}_async_plan.jsonl",
            repo_root / "data" / "Augmented" / "DAGs" / f"gaia_cat_{cat}_async_plan.jsonl",
        ]
        for candidate in candidates:
            if candidate.exists():
                return candidate
    return None


def _load_async_records(path: Path) -> Dict[str, Dict[str, Any]]:
    """Load async-reference JSONL records keyed by task id."""
    by_id: Dict[str, Dict[str, Any]] = {}
    for rec in load_jsonl(path):
        sample_id = _record_id(rec)
        if sample_id:
            by_id[sample_id] = rec
    return by_id


def _dependency_dag_from_async_record(
    async_record: Dict[str, Any],
    chain_dag: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Build the dependency-DAG reference while preserving the chain node labels."""
    base_dag = copy.deepcopy(chain_dag or (async_record.get("gold") or {}).get("plan_dag") or {})
    nodes = base_dag.get("nodes") or []
    sorted_nodes = sorted(
        enumerate(nodes),
        key=lambda item: (item[1].get("step_index", item[0]), item[0]),
    )
    step_to_node_id: Dict[int, str] = {}
    for local_idx, (_, node) in enumerate(sorted_nodes):
        node_id = node.get("node_id", f"n{local_idx}")
        step_to_node_id[local_idx] = node_id
        try:
            step_to_node_id[int(node.get("step_index"))] = node_id
        except (TypeError, ValueError):
            pass

    edges: List[Dict[str, Any]] = []
    seen_edges = set()

    def add_dependency_edge(source_step: int, target_step: int, edge_types: Any) -> None:
        if isinstance(edge_types, str):
            edge_type = edge_types
        else:
            edge_type = "+".join(edge_types or []) if edge_types else "dependency"
        source = step_to_node_id.get(source_step, f"n{source_step}")
        target = step_to_node_id.get(target_step, f"n{target_step}")
        edge_key = (source, target, edge_type)
        if edge_key in seen_edges:
            return
        seen_edges.add(edge_key)
        edges.append({
            "source": source,
            "target": target,
            "edge_type": edge_type,
        })

    dependency_analysis = async_record.get("dependency_analysis") or {}

    for edge in dependency_analysis.get("edges") or []:
        if edge.get("from") is None or edge.get("to") is None:
            continue
        try:
            source_step = int(edge["from"])
            target_step = int(edge["to"])
        except (TypeError, ValueError):
            continue
        add_dependency_edge(source_step, target_step, edge.get("types") or [])

    for dep in dependency_analysis.get("dependencies") or []:
        if dep.get("step_id") is None:
            continue
        try:
            target_step = int(dep["step_id"])
        except (TypeError, ValueError):
            continue
        for parent in dep.get("parents") or []:
            if parent.get("parent_id") is None:
                continue
            try:
                source_step = int(parent["parent_id"])
            except (TypeError, ValueError):
                continue
            add_dependency_edge(source_step, target_step, parent.get("types") or [])
    return {"nodes": nodes, "edges": edges}


@dataclass
class PerSampleResult:
    """Per-sample evaluation result for CSV output."""
    sample_id: str
    dataset: str
    subset: str
    plan_type: str

    # Plan metrics
    node_f1: float  # Strict semantic one-to-one NodeF1
    span_node_f1: float
    strict_node_f1: float
    edge_f1: float  # Paper-facing EdgeF1; alias of semantic_edge_f1 in new outputs
    raw_edge_f1: float
    semantic_edge_f1: float  # Backward-compatible alias for edge_f1
    dw_order_f1: float
    span_reanchored_dw_order_f1: float
    index_dw_order_f1: float
    dw_order_precision: float
    dw_order_recall: float
    index_dw_order_precision: float
    index_dw_order_recall: float
    planning_score: float
    order_precision: float
    order_recall: float
    order_f1: float
    node_label_similarity: float
    ssi: float
    gold_node_count: int
    pred_node_count: int
    reference_selected: str
    chain_only_planning_score: Optional[float]
    chain_only_node_f1: Optional[float]
    chain_only_dw_order_f1: Optional[float]
    chain_only_semantic_edge_f1: Optional[float]
    chain_only_raw_edge_f1: Optional[float]
    chain_only_node_label_similarity: Optional[float]
    dependency_dag_planning_score: Optional[float]
    dependency_dag_node_f1: Optional[float]
    dependency_dag_dw_order_f1: Optional[float]
    dependency_dag_semantic_edge_f1: Optional[float]
    dependency_dag_raw_edge_f1: Optional[float]
    dependency_dag_node_label_similarity: Optional[float]
    augmented_best_planning_score: Optional[float]
    augmented_best_node_f1: Optional[float]
    augmented_best_dw_order_f1: Optional[float]
    augmented_best_semantic_edge_f1: Optional[float]
    augmented_best_raw_edge_f1: Optional[float]
    augmented_best_node_label_similarity: Optional[float]

    # Tool metrics (value metrics can be None for datasets without arguments)
    tool_name_f1: float
    param_name_f1: Optional[float]  # None = N/A (dataset has no arguments)
    type_aware_value_f1: Optional[float]  # Main value metric used in ToolUsageScore
    strict_type_aware_value_f1: Optional[float]
    normalized_type_aware_value_f1: Optional[float]
    tool_usage_score: Optional[float]
    gt_aligned_tool_success_rate: Optional[float]
    pred_exec_success_rate: Optional[float]
    observation_support_rate: Optional[float]
    has_arguments: bool
    gold_tool_count: int
    pred_tool_count: int

    # Answer metrics (optional)
    has_answer: bool
    exact_match: Optional[float]
    token_f1: Optional[float]
    alias_match: Optional[float]
    llm_judge_score: Optional[float]

    # Metadata
    parse_error: bool


def _mean_tool_usage(tool_scores: Dict[str, Any]) -> Optional[float]:
    values = [
        tool_scores.get("tool_name_f1"),
        tool_scores.get("param_name_f1"),
        tool_scores.get("type_aware_value_f1"),
    ]
    valid = [v for v in values if v is not None]
    if not valid:
        return None
    return sum(valid) / len(valid)


def _plan_node_f1(plan_scores: Dict[str, Any]) -> float:
    return float(plan_scores.get("strict_node_f1", plan_scores.get("node_f1", 0.0)) or 0.0)


def _plan_span_node_f1(plan_scores: Dict[str, Any]) -> float:
    return float(plan_scores.get("span_node_f1", 0.0) or 0.0)


def _plan_dw_order_f1(plan_scores: Dict[str, Any]) -> float:
    return float(plan_scores.get("dw_order_f1", 0.0) or 0.0)


def _plan_edge_f1(plan_scores: Dict[str, Any]) -> float:
    return float(plan_scores.get("semantic_edge_f1", plan_scores.get("edge_f1", 0.0)) or 0.0)


def _plan_raw_edge_f1(plan_scores: Dict[str, Any]) -> float:
    return float(plan_scores.get("raw_edge_f1", plan_scores.get("edge_f1", 0.0)) or 0.0)


def _planning_score_from_components(plan_scores: Dict[str, Any]) -> float:
    return (_plan_node_f1(plan_scores) + _plan_edge_f1(plan_scores)) / 2.0


def _is_success_output(output: Any) -> bool:
    text = str(output or "")
    if "Output:\n" in text:
        text = text.split("Output:\n", 1)[1]
    norm = text.strip().lower()
    if not norm:
        return False
    return not (
        norm.startswith("[error]")
        or norm.startswith("error:")
        or norm.startswith("execution failed:")
    )


def _is_scored_tool_call(call: Dict[str, Any]) -> bool:
    """Return False for administrative calls that should not affect tool metrics."""
    return _extract_tool_name(call.get("tool_id", "") or "") not in {"submit_final_answer"}


def _extract_output_text(output: Any) -> str:
    text = str(output or "")
    if "Output:\n" in text:
        return text.split("Output:\n", 1)[1]
    return text


def _extract_answer_string(answer_obj: Any) -> str:
    """Extract the most human-readable answer string from nested answer payloads."""
    if isinstance(answer_obj, dict):
        for key in ("answer", "final_answer", "response", "output", "result"):
            if key in answer_obj:
                return _extract_answer_string(answer_obj.get(key))
        try:
            return json.dumps(answer_obj, ensure_ascii=False)
        except Exception:
            return str(answer_obj)
    if isinstance(answer_obj, list):
        try:
            return json.dumps(answer_obj, ensure_ascii=False)
        except Exception:
            return str(answer_obj)
    if answer_obj is None:
        return ""
    return str(answer_obj)


def _extract_submit_answer_from_tool_calls(tool_calls: List[Dict[str, Any]]) -> Tuple[str, bool]:
    """Extract the latest submit_final_answer payload from executed tool calls."""
    for call in reversed(tool_calls or []):
        if not isinstance(call, dict) or call.get("tool_id") != "submit_final_answer":
            continue
        for arg in call.get("arguments", []) or []:
            if isinstance(arg, dict) and arg.get("name") == "answer":
                return _extract_answer_string(arg.get("value")), True
        return "", True
    return "", False


def _extract_effective_submit_info(pred: Dict[str, Any]) -> Tuple[str, str, str, str]:
    """
    Resolve the effective final handoff answer used for answer-review diagnostics.

    Preference order:
      1. Native submit_final_answer tool call
      2. Explicit pred.submit_answer
      3. _answer_handoff.effective_submit_answer
      4. Missing
    """
    tool_calls = pred.get("tool_calls") or []
    handoff = pred.get("_answer_handoff", {}) or {}
    tool_submit_answer, submit_tool_present = _extract_submit_answer_from_tool_calls(tool_calls)
    if str(tool_submit_answer or "").strip():
        return (
            tool_submit_answer,
            str(handoff.get("source") or "native_submit_tool"),
            str(handoff.get("status") or "native_submit"),
            str(handoff.get("reason") or "Stage 4 submit_final_answer provided a non-empty answer."),
        )
    explicit_submit_answer = _extract_answer_string(pred.get("submit_answer"))
    if str(explicit_submit_answer or "").strip():
        return (
            explicit_submit_answer,
            str(handoff.get("source") or "pred_submit_answer_field"),
            str(handoff.get("status") or ("submit_tool_empty" if submit_tool_present else "stage5_backfill")),
            str(
                handoff.get("reason")
                or (
                    "A submit_final_answer call was present but empty; using the normalized submit_answer field."
                    if submit_tool_present
                    else "Using the normalized submit_answer field as the effective final handoff."
                )
            ),
        )

    handoff_submit_answer = _extract_answer_string(handoff.get("effective_submit_answer"))
    if str(handoff_submit_answer or "").strip():
        return (
            handoff_submit_answer,
            str(handoff.get("source") or "answer_handoff_fallback"),
            str(handoff.get("status") or ("submit_tool_empty" if submit_tool_present else "stage5_backfill")),
            str(
                handoff.get("reason")
                or (
                    "A submit_final_answer call was present but empty; using _answer_handoff.effective_submit_answer."
                    if submit_tool_present
                    else "Using _answer_handoff.effective_submit_answer as the effective final handoff."
                )
            ),
        )

    final_answer_fallback = _extract_answer_string(pred.get("final_answer"))
    if str(final_answer_fallback or "").strip():
        return (
            final_answer_fallback,
            str(handoff.get("source") or "legacy_final_answer_fallback"),
            str(handoff.get("status") or ("legacy_submit_tool_empty_backfill" if submit_tool_present else "legacy_missing_submit_tool_backfill")),
            str(
                handoff.get("reason")
                or (
                    "Legacy prediction contains an empty submit_final_answer call; using pred.final_answer as the effective final handoff."
                    if submit_tool_present
                    else "Legacy prediction lacks normalized handoff metadata; using pred.final_answer as the effective final handoff."
                )
            ),
        )

    if submit_tool_present:
        return "", "native_submit_tool", "submit_tool_empty", "Stage 4 submit_final_answer was called, but its answer argument was empty."
    return "", "missing", str(handoff.get("status") or "missing"), str(handoff.get("reason") or "No effective final handoff answer was available.")


def _infer_model_name(path: Path, meta: Dict[str, Any]) -> str:
    model_name = meta.get("model_name")
    if model_name:
        return str(model_name)
    if path.stem.startswith("unified."):
        return path.stem[len("unified."):]
    return "unknown_model"


def _shorten_sample_id(sample_id: str) -> str:
    sample_id = str(sample_id or "")
    if len(sample_id) <= 8:
        return sample_id
    return f"{sample_id[:8]}..."


def _shorten_model_name(model_name: str) -> str:
    model_name = str(model_name or "").strip()
    explicit_aliases = {
        "Llama-4-Maverick-17B-128E-Instruct-FP8": "Llama-4",
        "Llama-3.3-70B-Instruct": "Llama-3.3-70B",
        "Google-Gemma-3-27B": "Gemma-3-27B",
        "gemma-4-31B-it": "gemma-4-31B",
        "gemma-3-12b-it": "gemma-3-12B",
        "Mistral-Large-3-675B-Instruct-2512": "Mistral-Large",
        "Mistral-Small-3.2-24B-Instruct-2506": "Mistral-Small",
        "gpt-oss-120b": "gpt-oss-120b",
        "gpt-oss-20b": "gpt-oss-20b",
        "NVIDIA-Nemotron-3-Super-120B-A12B": "Nemotron-120B",
        "Microsoft-Phi-4-multimodal-instruct": "Phi-4-mm",
        "Llama-3.2-90B-Vision-Instruct": "Llama-3.2-90B-V",
    }
    if model_name in explicit_aliases:
        return explicit_aliases[model_name]

    compact = model_name
    compact = compact.replace("Google-", "")
    compact = compact.replace("-Instruct-FP8", "")
    compact = compact.replace("-Instruct", "")
    compact = compact.replace("-multimodal-instruct", "-mm")
    compact = compact.replace("-it", "")
    compact = compact.replace("-Chat", "")
    compact = compact.replace("-preview", "")
    compact = compact.replace("-2506", "")
    compact = compact.replace("-2512", "")
    return compact


def _review_text_equivalent(a: str, b: str, evaluator: ASTEvaluationSystem) -> bool:
    a = str(a or "").strip()
    b = str(b or "").strip()
    if not a or not b:
        return False
    return a == b or evaluator._normalized_exact_match(a, b)


def _review_answer_is_correct(
    gold_answer_obj: Dict[str, Any],
    pred_answer: str,
    evaluator: ASTEvaluationSystem,
) -> bool:
    pred_answer = str(pred_answer or "").strip()
    if not pred_answer:
        return False
    score = evaluator.evaluate_answer(gold_answer_obj, {"answer": pred_answer})
    return bool(score.has_answer and (score.exact_match or 0.0) >= 1.0)


def _classify_answer_review_case(
    final_correct: bool,
    submit_correct: bool,
    submit_missing: bool,
    final_submit_same: bool,
) -> str:
    if final_correct and submit_correct:
        return "both_correct_same" if final_submit_same else "both_correct_changed"
    if final_correct and not submit_correct:
        return "final_correct_submit_missing" if submit_missing else "final_correct_submit_wrong"
    if not final_correct and submit_correct:
        return "final_wrong_submit_correct"
    if submit_missing:
        return "submit_missing"
    return "same_wrong" if final_submit_same else "changed_but_wrong"


def _build_answer_review_rows(
    path: Path,
    results: List[PerSampleResult],
    gold_records_by_id: Optional[Dict[str, Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    """Create lightweight answer review rows for manual inspection."""
    metrics_by_id = {r.sample_id: r for r in results}
    review_rows: List[Dict[str, Any]] = []
    review_evaluator = ASTEvaluationSystem(use_embeddings=False, verifier_model=None)

    for rec in load_jsonl(path):
        meta = dict(rec.get("meta", {}) or {})
        sample_id = str(meta.get("id", ""))
        gold = rec.get("gold") or {}
        pred = rec.get("pred") or rec.get("prediction") or {}
        if gold_records_by_id and sample_id in gold_records_by_id:
            canonical_rec = gold_records_by_id[sample_id]
            gold = canonical_rec.get("gold") or gold
            canonical_meta = canonical_rec.get("meta", {}) or {}
            for key in ("dataset", "subset", "split", "plan_type"):
                if key in canonical_meta:
                    meta[key] = canonical_meta[key]

        metric_row = metrics_by_id.get(sample_id)
        final_answer_obj = pred.get("final_answer", {}) or {}
        final_answer = _extract_answer_string(final_answer_obj)
        submit_answer, submit_source, handoff_status, handoff_reason = _extract_effective_submit_info(pred)

        gold_answer_obj = gold.get("final_answer", {}) or {}
        gold_answer = _extract_answer_string(gold_answer_obj.get("answer", gold_answer_obj))
        gold_aliases = gold_answer_obj.get("aliases", []) if isinstance(gold_answer_obj, dict) else []
        gold_tolerance = gold_answer_obj.get("tolerance") if isinstance(gold_answer_obj, dict) else None
        final_correct = _review_answer_is_correct(gold_answer_obj, final_answer, review_evaluator)
        submit_correct = _review_answer_is_correct(gold_answer_obj, submit_answer, review_evaluator)
        submit_missing = not str(submit_answer or "").strip()
        final_submit_same = _review_text_equivalent(final_answer, submit_answer, review_evaluator)
        auto_case = _classify_answer_review_case(
            final_correct=final_correct,
            submit_correct=submit_correct,
            submit_missing=submit_missing,
            final_submit_same=final_submit_same,
        )

        review_rows.append({
            "sample_id": _shorten_sample_id(sample_id),
            "model_name": _shorten_model_name(_infer_model_name(path, meta)),
            "dataset": meta.get("dataset", "unknown"),
            "subset": meta.get("subset", "unknown"),
            "gold_answer": gold_answer,
            "Human Eval": "",
            "pre_final_answer": final_answer,
            "submit_final_answer": submit_answer,
            "pred_final_answer": final_answer,
            "pred_submit_answer": submit_answer,
            "submit_source": submit_source,
            "answer_handoff_status": handoff_status,
            "answer_handoff_reason": handoff_reason,
            "Final Match": "y" if final_correct else "n",
            "Submit Match": "y" if submit_correct else "n",
            "Auto Case": auto_case,
            "gold_aliases": json.dumps(gold_aliases, ensure_ascii=False) if gold_aliases else "",
            "gold_tolerance": gold_tolerance if gold_tolerance is not None else "",
            "exact_match": metric_row.exact_match if metric_row else "",
            "token_f1": metric_row.token_f1 if metric_row else "",
            "alias_match": metric_row.alias_match if metric_row else "",
            "llm_judge_score": metric_row.llm_judge_score if metric_row else "",
            "parse_error": metric_row.parse_error if metric_row else "",
        })

    return review_rows


def _write_answer_review_csv(rows: List[Dict[str, Any]], path: Path) -> None:
    """Write manual-review answer CSV with stable field order."""
    if not rows:
        return
    fieldnames = _answer_review_fieldnames()
    with path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _answer_review_fieldnames() -> List[str]:
    return [
        "sample_id",
        "model_name",
        "dataset",
        "subset",
        "gold_answer",
        "Human Eval",
        "submit_final_answer",
        "pre_final_answer",
        "pred_final_answer",
        "pred_submit_answer",
        "submit_source",
        "answer_handoff_status",
        "answer_handoff_reason",
        "Final Match",
        "Submit Match",
        "Auto Case",
        "gold_aliases",
        "gold_tolerance",
        "exact_match",
        "token_f1",
        "alias_match",
        "llm_judge_score",
        "parse_error",
    ]


def _write_answer_review_xlsx(rows: List[Dict[str, Any]], path: Path) -> None:
    """Write manual-review answer spreadsheet with wrapped text and frozen header."""
    if not rows:
        return

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fieldnames = _answer_review_fieldnames()
    wb = Workbook()
    ws = wb.active
    ws.title = "answer_review"

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    green_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    orange_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    red_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    blue_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")

    ws.append(fieldnames)
    for col_idx, field in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment

    for row in rows:
        ws.append([row.get(field, "") for field in fieldnames])

    widths = {
        "sample_id": 38,
        "model_name": 34,
        "dataset": 12,
        "subset": 16,
        "gold_answer": 48,
        "Human Eval": 12,
        "pre_final_answer": 48,
        "submit_final_answer": 48,
        "pred_final_answer": 48,
        "pred_submit_answer": 48,
        "submit_source": 22,
        "answer_handoff_status": 30,
        "answer_handoff_reason": 54,
        "Final Match": 12,
        "Submit Match": 12,
        "Auto Case": 30,
        "gold_aliases": 28,
        "gold_tolerance": 14,
        "exact_match": 12,
        "token_f1": 12,
        "alias_match": 12,
        "llm_judge_score": 14,
        "parse_error": 12,
    }

    for idx, field in enumerate(fieldnames, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(field, 20)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap_alignment

    col_idx = {field: idx + 1 for idx, field in enumerate(fieldnames)}
    auto_case_fills = {
        "both_correct_same": green_fill,
        "both_correct_changed": green_fill,
        "final_correct_submit_wrong": yellow_fill,
        "final_correct_submit_missing": yellow_fill,
        "final_wrong_submit_correct": blue_fill,
        "submit_missing": orange_fill,
        "changed_but_wrong": orange_fill,
        "same_wrong": red_fill,
    }

    for row_idx in range(2, ws.max_row + 1):
        final_match = str(ws.cell(row=row_idx, column=col_idx["Final Match"]).value or "").strip().lower()
        submit_match = str(ws.cell(row=row_idx, column=col_idx["Submit Match"]).value or "").strip().lower()
        auto_case = str(ws.cell(row=row_idx, column=col_idx["Auto Case"]).value or "").strip()

        if final_match == "y":
            ws.cell(row=row_idx, column=col_idx["pre_final_answer"]).fill = green_fill
            ws.cell(row=row_idx, column=col_idx["pred_final_answer"]).fill = green_fill
            ws.cell(row=row_idx, column=col_idx["Final Match"]).fill = green_fill
        elif final_match == "n":
            ws.cell(row=row_idx, column=col_idx["Final Match"]).fill = red_fill

        if submit_match == "y":
            ws.cell(row=row_idx, column=col_idx["submit_final_answer"]).fill = green_fill
            ws.cell(row=row_idx, column=col_idx["pred_submit_answer"]).fill = green_fill
            ws.cell(row=row_idx, column=col_idx["Submit Match"]).fill = green_fill
        elif submit_match == "n":
            ws.cell(row=row_idx, column=col_idx["Submit Match"]).fill = red_fill

        fill = auto_case_fills.get(auto_case)
        if fill is not None:
            ws.cell(row=row_idx, column=col_idx["Auto Case"]).fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    from openpyxl.worksheet.datavalidation import DataValidation

    human_eval_col = fieldnames.index("Human Eval") + 1
    human_eval_letter = get_column_letter(human_eval_col)
    human_eval_validation = DataValidation(type="list", formula1='"y,n"', allow_blank=True)
    human_eval_validation.prompt = "Write y if you judge the answer correct; n otherwise."
    human_eval_validation.error = "Use y or n."
    ws.add_data_validation(human_eval_validation)
    if ws.max_row >= 2:
        human_eval_validation.add(f"{human_eval_letter}2:{human_eval_letter}{ws.max_row}")

    wb.save(path)


def _load_existing_human_eval(path: Path) -> Dict[Tuple[str, str], str]:
    """Load existing manual review labels from an answer review CSV/XLSX file."""
    labels: Dict[Tuple[str, str], str] = {}
    if not path.exists():
        return labels

    if path.suffix.lower() == ".csv":
        with path.open(newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                key = (str(row.get("sample_id", "")).strip(), str(row.get("model_name", "")).strip())
                value = str(row.get("Human Eval", "")).strip()
                if key != ("", "") and value:
                    labels[key] = value
        return labels

    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True)
        ws = wb["answer_review"] if "answer_review" in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return labels
        index = {str(name): i for i, name in enumerate(header)}
        if "Human Eval" not in index:
            return labels
        for row in rows:
            if not row:
                continue
            sample_id = str(row[index.get("sample_id", -1)] or "").strip()
            model_name = str(row[index.get("model_name", -1)] or "").strip()
            value = str(row[index["Human Eval"]] or "").strip()
            if sample_id and model_name and value:
                labels[(sample_id, model_name)] = value
    return labels


def _preserve_existing_human_eval(
    rows: List[Dict[str, Any]],
    review_csv_path: Path,
    review_xlsx_path: Path,
) -> None:
    """Carry forward existing manual y/n labels when regenerating review files."""
    existing_labels: Dict[Tuple[str, str], str] = {}
    existing_labels.update(_load_existing_human_eval(review_csv_path))
    existing_labels.update(_load_existing_human_eval(review_xlsx_path))
    if not existing_labels:
        return

    for row in rows:
        key = (str(row.get("sample_id", "")).strip(), str(row.get("model_name", "")).strip())
        if key in existing_labels:
            row["Human Eval"] = existing_labels[key]


def evaluate_file(
    path: Path,
    use_embeddings: bool = True,
    gold_records_by_id: Optional[Dict[str, Dict[str, Any]]] = None,
    gaia_async_records_by_id: Optional[Dict[str, Dict[str, Any]]] = None,
    gaia_reference_mode: str = "auto",
    verifier_model: Optional[str] = None,
    plan_source: str = "stage1",
) -> Tuple[List[PerSampleResult], Dict[str, Any]]:
    """
    Evaluate a single file with gold + pred.

    Returns:
        Tuple of (per_sample_results, summary_dict)
    """
    evaluator = ASTEvaluationSystem(
        use_embeddings=use_embeddings,
        verifier_model=verifier_model,
    )

    results: List[PerSampleResult] = []
    buckets: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)

    for rec in load_jsonl(path):
        gold = rec.get("gold") or {}
        pred = rec.get("pred") or rec.get("prediction") or {}
        meta = dict(rec.get("meta", {}) or {})
        meta["plan_source"] = plan_source
        query = rec.get("query", {}) or {}

        sample_id = str(meta.get("id", ""))
        canonical_tool_env = None
        if gold_records_by_id and sample_id in gold_records_by_id:
            canonical_rec = gold_records_by_id[sample_id]
            gold = canonical_rec.get("gold") or gold
            canonical_tool_env = canonical_rec.get("tool_environment")
            canonical_meta = canonical_rec.get("meta", {}) or {}
            for key in ("dataset", "subset", "split", "plan_type", "has_arguments", "has_answer"):
                if key in canonical_meta:
                    meta[key] = canonical_meta[key]

        dataset_for_ref = str(meta.get("dataset", "") or "").lower()
        if (
            gaia_reference_mode in {"auto", "augmented", "both"}
            and dataset_for_ref == "gaia"
            and gaia_async_records_by_id
            and sample_id in gaia_async_records_by_id
        ):
            chain_dag = gold.get("plan_dag") or gold.get("canonical_plan_dag") or {}
            dependency_dag = _dependency_dag_from_async_record(
                gaia_async_records_by_id[sample_id],
                chain_dag=chain_dag,
            )
            if dependency_dag.get("nodes") and dependency_dag.get("edges"):
                meta["reference_plan_dags"] = [
                    {"name": "chain", "dag": chain_dag},
                    {"name": "dag", "dag": dependency_dag},
                ]

        attachments = query.get("attachments", []) or []
        meta["has_image_attachment"] = any(
            str(att.get("file_name", "")).lower().endswith((".png", ".jpg", ".jpeg", ".webp", ".gif"))
            for att in attachments
        )
        if "supports_native_vision" not in meta:
            model_name = meta.get("model_name")
            if not model_name and path.stem.startswith("unified."):
                model_name = path.stem[len("unified."):]
                meta["model_name"] = model_name
            model_name_norm = str(model_name or "").lower()
            meta["supports_native_vision"] = bool(
                model_name_norm and (
                    "vision" in model_name_norm
                    or "multimodal" in model_name_norm
                    or "phi-4" in model_name_norm
                )
            )

        # Check for parse error
        parse_error = pred.get("_parse_error", False) or "_error" in pred

        # Auto-detect has_arguments if not in meta
        if "has_arguments" not in meta:
            gold_calls = gold.get("tool_calls") or []
            has_args = False
            for c in gold_calls:
                args = c.get("arguments", [])
                if args:
                    if isinstance(args, list) and len(args) > 0:
                        has_args = True
                        break
                    elif isinstance(args, dict) and len(args) > 0:
                        has_args = True
                        break
            meta["has_arguments"] = has_args

        # Pass available tools for filtering invented tools (Option C)
        # This ensures LLM-invented tools don't affect evaluation metrics
        tool_env = canonical_tool_env if canonical_tool_env is not None else rec.get("tool_environment", {})
        _tools = normalize_tool_environment(tool_env)
        if "available_tools" not in meta and _tools:
            meta["available_tools"] = _tools

        # Evaluate with metadata for dataset-aware handling
        scores = evaluator.evaluate_record(gold, pred, meta, query=query)

        dataset = meta.get("dataset", "unknown")
        subset = meta.get("subset", "unknown")
        sample_id = str(meta.get("id", ""))
        plan_type = meta.get("plan_type", "unknown")

        pred_calls = pred.get("tool_calls") or []
        pred_outputs = pred.get("tool_outputs") or []
        pred_call_outputs = []
        for idx, call in enumerate(pred_calls):
            output = pred_outputs[idx] if idx < len(pred_outputs) else None
            if _is_scored_tool_call(call):
                pred_call_outputs.append((call, output))

        scored_pred_calls = [call for call, _ in pred_call_outputs]
        scored_pred_outputs = [output for _, output in pred_call_outputs]
        success_flags = [_is_success_output(output) for output in scored_pred_outputs]
        if len(success_flags) < len(scored_pred_calls):
            success_flags.extend([False] * (len(scored_pred_calls) - len(success_flags)))
        success_flags = success_flags[:len(scored_pred_calls)]

        pred_exec_success_rate = (
            sum(1 for ok in success_flags if ok) / len(scored_pred_calls)
            if scored_pred_calls else None
        )

        gold_tools = [
            _extract_tool_name(call.get("tool_id", ""))
            for call in (gold.get("tool_calls") or [])
            if call.get("tool_id") and _is_scored_tool_call(call)
        ]
        successful_pred_tools = [
            _extract_tool_name(call.get("tool_id", ""))
            for call, ok in zip(scored_pred_calls, success_flags)
            if ok and call.get("tool_id")
        ]
        gold_counter = Counter(gold_tools)
        pred_counter = Counter(successful_pred_tools)
        matched_success = sum(min(gold_counter[key], pred_counter[key]) for key in gold_counter)
        gt_aligned_success_rate = (
            matched_success / len(gold_tools) if gold_tools else None
        )

        successful_observation_text = "\n\n".join(
            _extract_output_text(output)
            for output, ok in zip(scored_pred_outputs, success_flags)
            if ok
        )
        observation_support_rate = None
        if scores["answer"]["has_answer"]:
            observation_support_rate = 0.0
            if successful_observation_text.strip():
                observation_answer = evaluator.evaluate_answer(
                    gold.get("final_answer", {}) or {},
                    {"answer": successful_observation_text},
                    query_text=str(query.get("user_query", "") or ""),
                )
                observation_support_rate = float(observation_answer.exact_match or 0.0)

        # Create per-sample result
        plan_node_f1 = _plan_node_f1(scores["plan"])
        plan_span_node_f1 = scores["plan"].get("span_node_f1", 0.0)
        plan_strict_node_f1 = scores["plan"].get("strict_node_f1", scores["plan"]["node_f1"])
        plan_edge_f1 = _plan_edge_f1(scores["plan"])
        plan_raw_edge_f1 = _plan_raw_edge_f1(scores["plan"])
        plan_dw_order_f1 = scores["plan"].get("dw_order_f1", 0.0)
        plan_planning_score = _planning_score_from_components(scores["plan"])
        result = PerSampleResult(
            sample_id=sample_id,
            dataset=dataset,
            subset=subset,
            plan_type=plan_type,

            node_f1=plan_node_f1,
            span_node_f1=plan_span_node_f1,
            strict_node_f1=plan_strict_node_f1,
            edge_f1=plan_edge_f1,
            raw_edge_f1=plan_raw_edge_f1,
            semantic_edge_f1=scores["plan"].get("semantic_edge_f1", plan_edge_f1),
            dw_order_f1=plan_dw_order_f1,
            span_reanchored_dw_order_f1=scores["plan"].get("span_reanchored_dw_order_f1", plan_dw_order_f1),
            index_dw_order_f1=scores["plan"].get("index_dw_order_f1", plan_dw_order_f1),
            dw_order_precision=scores["plan"].get("dw_order_precision", 0.0),
            dw_order_recall=scores["plan"].get("dw_order_recall", 0.0),
            index_dw_order_precision=scores["plan"].get("index_dw_order_precision", 0.0),
            index_dw_order_recall=scores["plan"].get("index_dw_order_recall", 0.0),
            planning_score=plan_planning_score,
            order_precision=scores["plan"].get("order_precision", 0.0),
            order_recall=scores["plan"].get("order_recall", 0.0),
            order_f1=scores["plan"].get("order_f1", 0.0),
            node_label_similarity=scores["plan"]["node_label_similarity"],
            ssi=scores["plan"]["ssi"],
            gold_node_count=scores["plan"]["gold_node_count"],
            pred_node_count=scores["plan"]["pred_node_count"],
            reference_selected=str(scores["plan"].get("reference_selected") or "chain"),
            chain_only_planning_score=scores["plan"].get("chain_only_planning_score"),
            chain_only_node_f1=scores["plan"].get("chain_only_node_f1"),
            chain_only_dw_order_f1=scores["plan"].get("chain_only_dw_order_f1"),
            chain_only_semantic_edge_f1=scores["plan"].get("chain_only_semantic_edge_f1"),
            chain_only_raw_edge_f1=scores["plan"].get("chain_only_raw_edge_f1"),
            chain_only_node_label_similarity=scores["plan"].get("chain_only_node_label_similarity"),
            dependency_dag_planning_score=scores["plan"].get("dependency_dag_planning_score"),
            dependency_dag_node_f1=scores["plan"].get("dependency_dag_node_f1"),
            dependency_dag_dw_order_f1=scores["plan"].get("dependency_dag_dw_order_f1"),
            dependency_dag_semantic_edge_f1=scores["plan"].get("dependency_dag_semantic_edge_f1"),
            dependency_dag_raw_edge_f1=scores["plan"].get("dependency_dag_raw_edge_f1"),
            dependency_dag_node_label_similarity=scores["plan"].get("dependency_dag_node_label_similarity"),
            augmented_best_planning_score=scores["plan"].get("augmented_best_planning_score"),
            augmented_best_node_f1=scores["plan"].get("augmented_best_node_f1"),
            augmented_best_dw_order_f1=scores["plan"].get("augmented_best_dw_order_f1"),
            augmented_best_semantic_edge_f1=scores["plan"].get("augmented_best_semantic_edge_f1"),
            augmented_best_raw_edge_f1=scores["plan"].get("augmented_best_raw_edge_f1"),
            augmented_best_node_label_similarity=scores["plan"].get("augmented_best_node_label_similarity"),

            tool_name_f1=scores["tool"]["tool_name_f1"],
            param_name_f1=scores["tool"]["param_name_f1"],  # May be None
            type_aware_value_f1=scores["tool"].get("type_aware_value_f1"),
            strict_type_aware_value_f1=scores["tool"].get("strict_type_aware_value_f1"),
            normalized_type_aware_value_f1=scores["tool"].get("normalized_type_aware_value_f1"),
            tool_usage_score=_mean_tool_usage(scores["tool"]),
            gt_aligned_tool_success_rate=gt_aligned_success_rate,
            pred_exec_success_rate=pred_exec_success_rate,
            observation_support_rate=observation_support_rate,
            has_arguments=scores["tool"]["has_arguments"],
            gold_tool_count=scores["tool"]["gold_tool_count"],
            pred_tool_count=scores["tool"]["pred_tool_count"],

            has_answer=scores["answer"]["has_answer"],
            exact_match=scores["answer"]["exact_match"],
            token_f1=scores["answer"]["token_f1"],
            alias_match=scores["answer"].get("alias_match"),
            llm_judge_score=scores["answer"].get("llm_judge_score"),

            parse_error=parse_error,
        )
        results.append(result)

        # Add to bucket
        key = (dataset, subset)
        buckets[key].append({
            "scores": scores,
            "parse_error": parse_error,
            "has_arguments": scores["tool"]["has_arguments"],
        })

    # Compute summary
    summary: Dict[str, Any] = {}

    for (dataset, subset), items in buckets.items():
        n = len(items)
        parse_errors = sum(1 for it in items if it["parse_error"])

        score_list = [it["scores"] for it in items]
        has_args_list = [it["has_arguments"] for it in items]

        # Plan metrics (always available)
        plan_node_f1 = mean([_plan_node_f1(s["plan"]) for s in score_list])
        plan_span_node_f1 = mean([_plan_span_node_f1(s["plan"]) for s in score_list])
        plan_strict_node_f1 = mean([s["plan"].get("strict_node_f1", s["plan"]["node_f1"]) for s in score_list])
        plan_edge_f1 = mean([_plan_edge_f1(s["plan"]) for s in score_list])
        plan_raw_edge_f1 = mean([_plan_raw_edge_f1(s["plan"]) for s in score_list])
        plan_semantic_edge_f1 = mean([s["plan"].get("semantic_edge_f1", _plan_edge_f1(s["plan"])) for s in score_list])
        plan_dw_order_f1 = mean([s["plan"].get("dw_order_f1", 0.0) for s in score_list])
        plan_span_reanchored_dw_order_f1 = mean([s["plan"].get("span_reanchored_dw_order_f1", s["plan"].get("dw_order_f1", 0.0)) for s in score_list])
        plan_index_dw_order_f1 = mean([s["plan"].get("index_dw_order_f1", s["plan"].get("dw_order_f1", 0.0)) for s in score_list])
        plan_dw_order_precision = mean([s["plan"].get("dw_order_precision", 0.0) for s in score_list])
        plan_dw_order_recall = mean([s["plan"].get("dw_order_recall", 0.0) for s in score_list])
        plan_index_dw_order_precision = mean([s["plan"].get("index_dw_order_precision", 0.0) for s in score_list])
        plan_index_dw_order_recall = mean([s["plan"].get("index_dw_order_recall", 0.0) for s in score_list])
        plan_planning_score = mean([_planning_score_from_components(s["plan"]) for s in score_list])
        plan_order_precision = mean([s["plan"].get("order_precision", 0.0) for s in score_list])
        plan_order_recall = mean([s["plan"].get("order_recall", 0.0) for s in score_list])
        plan_order_f1 = mean([s["plan"].get("order_f1", 0.0) for s in score_list])
        plan_node_label_sim = mean([s["plan"]["node_label_similarity"] for s in score_list])
        plan_ssi = mean([s["plan"]["ssi"] for s in score_list])
        plan_reference_selected_counts = dict(Counter(
            str(s["plan"].get("reference_selected") or "chain")
            for s in score_list
        ))

        # Tool metrics
        # Only average Tool F1 over samples whose gold contains non-administrative tools.
        # This avoids mixing in pure-thought/direct-answer samples where the gold
        # intentionally has no executable tool usage beyond submit_final_answer.
        samples_with_gold_tools = [
            s for s in score_list
            if (s["tool"].get("gold_tool_count", 0) or 0) > 0
        ]
        tool_name_f1_all = mean([s["tool"]["tool_name_f1"] for s in score_list])
        tool_name_f1 = (
            mean([s["tool"]["tool_name_f1"] for s in samples_with_gold_tools])
            if samples_with_gold_tools else None
        )
        bucket_results = [r for r in results if r.dataset == dataset and r.subset == subset]
        tool_usage_score = mean([r.tool_usage_score for r in bucket_results if r.tool_usage_score is not None and r.gold_tool_count > 0])
        gt_aligned_tool_success = mean([r.gt_aligned_tool_success_rate for r in bucket_results if r.gt_aligned_tool_success_rate is not None and r.gold_tool_count > 0])
        pred_exec_success_rate = mean([r.pred_exec_success_rate for r in bucket_results if r.pred_exec_success_rate is not None and r.pred_tool_count > 0])
        observation_support_rate = mean([r.observation_support_rate for r in bucket_results if r.has_answer and r.observation_support_rate is not None])

        # Param metrics - only aggregate if at least some samples have arguments
        samples_with_args = [s for s, has_args in zip(score_list, has_args_list) if has_args]
        if samples_with_args:
            param_name_f1 = mean([s["tool"]["param_name_f1"] for s in samples_with_args])
            type_aware_value_f1 = mean([s["tool"].get("type_aware_value_f1") for s in samples_with_args])
            strict_type_aware_value_f1 = mean([s["tool"].get("strict_type_aware_value_f1") for s in samples_with_args])
            normalized_type_aware_value_f1 = mean([s["tool"].get("normalized_type_aware_value_f1") for s in samples_with_args])
        else:
            param_name_f1 = None  # N/A for this subset
            type_aware_value_f1 = None
            strict_type_aware_value_f1 = None
            normalized_type_aware_value_f1 = None

        # Answer metrics
        ans_scores = [s["answer"] for s in score_list if s["answer"]["has_answer"]]
        ans_em = mean([a["exact_match"] for a in ans_scores]) if ans_scores else None
        ans_f1 = mean([a["token_f1"] for a in ans_scores]) if ans_scores else None
        ans_alias = mean([a.get("alias_match") for a in ans_scores]) if ans_scores else None
        ans_judge = mean([a.get("llm_judge_score") for a in ans_scores]) if ans_scores else None

        summary_key = f"{dataset}/{subset}"
        summary[summary_key] = {
            "num_samples": n,
            "parse_errors": parse_errors,
            "parse_error_rate": parse_errors / n if n > 0 else 0.0,
            "tool_required_samples": len(samples_with_gold_tools),
            "has_argument_samples": len(samples_with_args),
            "plan": {
                "node_f1": plan_node_f1,
                "span_node_f1": plan_span_node_f1,
                "strict_node_f1": plan_strict_node_f1,
                "edge_f1": plan_edge_f1,
                "raw_edge_f1": plan_raw_edge_f1,
                "semantic_edge_f1": plan_semantic_edge_f1,
                "dw_order_f1": plan_dw_order_f1,
                "span_reanchored_dw_order_f1": plan_span_reanchored_dw_order_f1,
                "index_dw_order_f1": plan_index_dw_order_f1,
                "dw_order_precision": plan_dw_order_precision,
                "dw_order_recall": plan_dw_order_recall,
                "index_dw_order_precision": plan_index_dw_order_precision,
                "index_dw_order_recall": plan_index_dw_order_recall,
                "planning_score": plan_planning_score,
                "order_precision": plan_order_precision,
                "order_recall": plan_order_recall,
                "order_f1": plan_order_f1,
                "node_label_similarity": plan_node_label_sim,
                "ssi": plan_ssi,
                "reference_selected_counts": plan_reference_selected_counts,
                "chain_only_planning_score": mean([s["plan"].get("chain_only_planning_score") for s in score_list]),
                "chain_only_node_f1": mean([s["plan"].get("chain_only_node_f1") for s in score_list]),
                "chain_only_dw_order_f1": mean([s["plan"].get("chain_only_dw_order_f1") for s in score_list]),
                "chain_only_semantic_edge_f1": mean([s["plan"].get("chain_only_semantic_edge_f1") for s in score_list]),
                "chain_only_raw_edge_f1": mean([s["plan"].get("chain_only_raw_edge_f1") for s in score_list]),
                "chain_only_node_label_similarity": mean([s["plan"].get("chain_only_node_label_similarity") for s in score_list]),
                "dependency_dag_planning_score": mean([s["plan"].get("dependency_dag_planning_score") for s in score_list]),
                "dependency_dag_node_f1": mean([s["plan"].get("dependency_dag_node_f1") for s in score_list]),
                "dependency_dag_dw_order_f1": mean([s["plan"].get("dependency_dag_dw_order_f1") for s in score_list]),
                "dependency_dag_semantic_edge_f1": mean([s["plan"].get("dependency_dag_semantic_edge_f1") for s in score_list]),
                "dependency_dag_raw_edge_f1": mean([s["plan"].get("dependency_dag_raw_edge_f1") for s in score_list]),
                "dependency_dag_node_label_similarity": mean([s["plan"].get("dependency_dag_node_label_similarity") for s in score_list]),
                "augmented_best_planning_score": mean([s["plan"].get("augmented_best_planning_score") for s in score_list]),
                "augmented_best_node_f1": mean([s["plan"].get("augmented_best_node_f1") for s in score_list]),
                "augmented_best_dw_order_f1": mean([s["plan"].get("augmented_best_dw_order_f1") for s in score_list]),
                "augmented_best_semantic_edge_f1": mean([s["plan"].get("augmented_best_semantic_edge_f1") for s in score_list]),
                "augmented_best_raw_edge_f1": mean([s["plan"].get("augmented_best_raw_edge_f1") for s in score_list]),
                "augmented_best_node_label_similarity": mean([s["plan"].get("augmented_best_node_label_similarity") for s in score_list]),
            },
            "tool": {
                "tool_name_f1": tool_name_f1,
                "tool_name_f1_all": tool_name_f1_all,
                "has_tool_samples": len(samples_with_gold_tools) > 0,
                "param_name_f1": param_name_f1,  # None if N/A
                "type_aware_value_f1": type_aware_value_f1,
                "strict_type_aware_value_f1": strict_type_aware_value_f1,
                "normalized_type_aware_value_f1": normalized_type_aware_value_f1,
                "tool_usage_score": tool_usage_score,
                "gt_aligned_tool_success_rate": gt_aligned_tool_success,
                "pred_exec_success_rate": pred_exec_success_rate,
                "observation_support_rate": observation_support_rate,
                "has_arguments": len(samples_with_args) > 0,  # Explicit flag for visualization
            },
            "answer": {
                "num_with_answer": len(ans_scores),
                "exact_match": ans_em,
                "token_f1": ans_f1,
                "alias_match": ans_alias,
                "llm_judge_score": ans_judge,
            } if ans_scores else None,
        }

    # Compute overall summary
    all_items = [it for items in buckets.values() for it in items]
    all_scores = [it["scores"] for it in all_items]
    all_has_args = [it["has_arguments"] for it in all_items]
    all_parse_errors = sum(it["parse_error"] for it in all_items)
    total_n = len(all_scores)

    if total_n > 0:
        # Separate by argument availability
        with_args = [(s, h) for s, h in zip(all_scores, all_has_args) if h]
        without_args = [(s, h) for s, h in zip(all_scores, all_has_args) if not h]
        with_gold_tools = [
            s for s in all_scores
            if (s["tool"].get("gold_tool_count", 0) or 0) > 0
        ]
        result_with_gold_tools = [r for r in results if r.gold_tool_count > 0]
        result_with_pred_tools = [r for r in results if r.pred_tool_count > 0]

        all_ans = [s["answer"] for s in all_scores if s["answer"]["has_answer"]]

        summary["_overall"] = {
            "num_samples": total_n,
            "parse_errors": all_parse_errors,
            "parse_error_rate": all_parse_errors / total_n,
            "tool_required_samples": len(with_gold_tools),
            "samples_with_arguments": len(with_args),
            "samples_without_arguments": len(without_args),
            "plan": {
                "node_f1": mean([_plan_node_f1(s["plan"]) for s in all_scores]),
                "span_node_f1": mean([_plan_span_node_f1(s["plan"]) for s in all_scores]),
                "strict_node_f1": mean([s["plan"].get("strict_node_f1", s["plan"]["node_f1"]) for s in all_scores]),
                "edge_f1": mean([_plan_edge_f1(s["plan"]) for s in all_scores]),
                "raw_edge_f1": mean([_plan_raw_edge_f1(s["plan"]) for s in all_scores]),
                "semantic_edge_f1": mean([s["plan"].get("semantic_edge_f1", _plan_edge_f1(s["plan"])) for s in all_scores]),
                "dw_order_f1": mean([s["plan"].get("dw_order_f1", 0.0) for s in all_scores]),
                "span_reanchored_dw_order_f1": mean([s["plan"].get("span_reanchored_dw_order_f1", s["plan"].get("dw_order_f1", 0.0)) for s in all_scores]),
                "index_dw_order_f1": mean([s["plan"].get("index_dw_order_f1", s["plan"].get("dw_order_f1", 0.0)) for s in all_scores]),
                "dw_order_precision": mean([s["plan"].get("dw_order_precision", 0.0) for s in all_scores]),
                "dw_order_recall": mean([s["plan"].get("dw_order_recall", 0.0) for s in all_scores]),
                "index_dw_order_precision": mean([s["plan"].get("index_dw_order_precision", 0.0) for s in all_scores]),
                "index_dw_order_recall": mean([s["plan"].get("index_dw_order_recall", 0.0) for s in all_scores]),
                "planning_score": mean([_planning_score_from_components(s["plan"]) for s in all_scores]),
                "order_precision": mean([s["plan"].get("order_precision", 0.0) for s in all_scores]),
                "order_recall": mean([s["plan"].get("order_recall", 0.0) for s in all_scores]),
                "order_f1": mean([s["plan"].get("order_f1", 0.0) for s in all_scores]),
                "node_label_similarity": mean([s["plan"]["node_label_similarity"] for s in all_scores]),
                "ssi": mean([s["plan"]["ssi"] for s in all_scores]),
                "reference_selected_counts": dict(Counter(
                    str(s["plan"].get("reference_selected") or "chain")
                    for s in all_scores
                )),
                "chain_only_planning_score": mean([s["plan"].get("chain_only_planning_score") for s in all_scores]),
                "chain_only_node_f1": mean([s["plan"].get("chain_only_node_f1") for s in all_scores]),
                "chain_only_dw_order_f1": mean([s["plan"].get("chain_only_dw_order_f1") for s in all_scores]),
                "chain_only_semantic_edge_f1": mean([s["plan"].get("chain_only_semantic_edge_f1") for s in all_scores]),
                "chain_only_raw_edge_f1": mean([s["plan"].get("chain_only_raw_edge_f1") for s in all_scores]),
                "chain_only_node_label_similarity": mean([s["plan"].get("chain_only_node_label_similarity") for s in all_scores]),
                "dependency_dag_planning_score": mean([s["plan"].get("dependency_dag_planning_score") for s in all_scores]),
                "dependency_dag_node_f1": mean([s["plan"].get("dependency_dag_node_f1") for s in all_scores]),
                "dependency_dag_dw_order_f1": mean([s["plan"].get("dependency_dag_dw_order_f1") for s in all_scores]),
                "dependency_dag_semantic_edge_f1": mean([s["plan"].get("dependency_dag_semantic_edge_f1") for s in all_scores]),
                "dependency_dag_raw_edge_f1": mean([s["plan"].get("dependency_dag_raw_edge_f1") for s in all_scores]),
                "dependency_dag_node_label_similarity": mean([s["plan"].get("dependency_dag_node_label_similarity") for s in all_scores]),
                "augmented_best_planning_score": mean([s["plan"].get("augmented_best_planning_score") for s in all_scores]),
                "augmented_best_node_f1": mean([s["plan"].get("augmented_best_node_f1") for s in all_scores]),
                "augmented_best_dw_order_f1": mean([s["plan"].get("augmented_best_dw_order_f1") for s in all_scores]),
                "augmented_best_semantic_edge_f1": mean([s["plan"].get("augmented_best_semantic_edge_f1") for s in all_scores]),
                "augmented_best_raw_edge_f1": mean([s["plan"].get("augmented_best_raw_edge_f1") for s in all_scores]),
                "augmented_best_node_label_similarity": mean([s["plan"].get("augmented_best_node_label_similarity") for s in all_scores]),
            },
            "tool": {
                "tool_name_f1": mean([s["tool"]["tool_name_f1"] for s in with_gold_tools]) if with_gold_tools else None,
                "tool_name_f1_all": mean([s["tool"]["tool_name_f1"] for s in all_scores]),
                "has_tool_samples": len(with_gold_tools) > 0,
                # Only compute param metrics from samples with arguments
                "param_name_f1": mean([s["tool"]["param_name_f1"] for s, _ in with_args]) if with_args else None,
                "type_aware_value_f1": mean([s["tool"].get("type_aware_value_f1") for s, _ in with_args]) if with_args else None,
                "strict_type_aware_value_f1": mean([s["tool"].get("strict_type_aware_value_f1") for s, _ in with_args]) if with_args else None,
                "normalized_type_aware_value_f1": mean([s["tool"].get("normalized_type_aware_value_f1") for s, _ in with_args]) if with_args else None,
                "tool_usage_score": mean([r.tool_usage_score for r in result_with_gold_tools if r.tool_usage_score is not None]),
                "gt_aligned_tool_success_rate": mean([r.gt_aligned_tool_success_rate for r in result_with_gold_tools if r.gt_aligned_tool_success_rate is not None]),
                "pred_exec_success_rate": mean([r.pred_exec_success_rate for r in result_with_pred_tools if r.pred_exec_success_rate is not None]),
                "observation_support_rate": mean([r.observation_support_rate for r in results if r.has_answer and r.observation_support_rate is not None]),
            },
            "answer": {
                "num_with_answer": len(all_ans),
                "exact_match": mean([a["exact_match"] for a in all_ans]) if all_ans else None,
                "token_f1": mean([a["token_f1"] for a in all_ans]) if all_ans else None,
                "alias_match": mean([a.get("alias_match") for a in all_ans]) if all_ans else None,
                "llm_judge_score": mean([a.get("llm_judge_score") for a in all_ans]) if all_ans else None,
            } if all_ans else None,
        }

    return results, summary




def main():
    parser = argparse.ArgumentParser(
        description="Evaluate LLM planning predictions (V2 - Dataset-Aware).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Key Improvements in V2:
- Dataset-aware metrics (Delta vs TaskBench/UltraTool)
- Proper handling of N/A for param_name_f1 and type_aware_value_f1 when dataset has no arguments
- Separate aggregation for samples with/without arguments

Examples:
    # Basic evaluation
    python dual_facet_evaluation_runner_v2.py --input results.jsonl

    # With CSV and JSON output
    python dual_facet_evaluation_runner_v2.py \\
        --input results.jsonl \\
        --output_csv per_sample.csv \\
        --output_summary summary.json
        """
    )
    parser.add_argument("--input", type=Path, required=True,
                        help="Input JSONL file with gold + pred")
    parser.add_argument("--output_csv", type=Path, default=None,
                        help="Output CSV for per-sample results")
    parser.add_argument("--output_summary", type=Path, default=None,
                        help="Output JSON for summary statistics")
    parser.add_argument("--gold_dataset", type=Path, default=None,
                        help="Optional canonical dataset JSON/JSONL used to override embedded gold by meta.id")
    parser.add_argument("--gaia_reference_mode", type=str, default="auto",
                        choices=["auto", "chain", "augmented", "both"],
                        help="GAIA planning reference mode. auto/augmented/both use original chain and dependency DAG as candidate references when async refs are available; chain preserves legacy chain-only scoring.")
    parser.add_argument("--gaia_async_dataset", type=Path, default=None,
                        help="Optional GAIA async JSONL reference file. If omitted, inferred from --gold_dataset when possible.")
    parser.add_argument("--verifier_model", type=str, default=None,
                        help="Optional LLM-as-a-judge model for answer verification (falls back silently if unavailable)")
    parser.add_argument("--no_embeddings", action="store_true",
                        help="Disable embedding-based similarity (use string similarity)")
    parser.add_argument("--plan_source", type=str, default="stage1",
                        choices=["stage1", "stage3", "abs", "abstract", "abs_plan_dag"],
                        help="Prediction plan field to evaluate. Default stage1 reads pred.abs_plan_dag; stage3 reads pred.plan_dag.")
    parser.add_argument("--quiet", action="store_true",
                        help="Don't print summary to console")
    parser.add_argument("--write_review_xlsx", action="store_true",
                        help="Also write answer_review XLSX files. By default, only CSV review files are produced.")

    args = parser.parse_args()

    print(f"[INFO] Evaluating: {args.input}")
    print(f"[INFO] Plan source: {args.plan_source}")

    gold_records_by_id = None
    if args.gold_dataset:
        gold_records_by_id = _load_gold_records(args.gold_dataset)
        print(f"[INFO] Loaded canonical gold records from: {args.gold_dataset} ({len(gold_records_by_id)} records)")

    gaia_async_records_by_id = None
    if args.gaia_reference_mode in {"auto", "augmented", "both"}:
        async_path = args.gaia_async_dataset or _infer_gaia_async_path(args.gold_dataset)
        if async_path and async_path.exists():
            gaia_async_records_by_id = _load_async_records(async_path)
            print(f"[INFO] Loaded GAIA async references from: {async_path} ({len(gaia_async_records_by_id)} records)")
        elif args.gaia_reference_mode in {"augmented", "both"}:
            raise FileNotFoundError(
                "GAIA augmented reference mode requested, but no async reference file was found. "
                "Pass --gaia_async_dataset explicitly or use --gaia_reference_mode chain."
            )

    results, summary = evaluate_file(
        args.input,
        use_embeddings=not args.no_embeddings,
        gold_records_by_id=gold_records_by_id,
        gaia_async_records_by_id=gaia_async_records_by_id,
        gaia_reference_mode=args.gaia_reference_mode,
        verifier_model=args.verifier_model,
        plan_source=args.plan_source,
    )

    if args.output_csv:
        write_csv(results, args.output_csv, PerSampleResult)
        print(f"[INFO] Per-sample results written to: {args.output_csv}")
        answer_review_rows = _build_answer_review_rows(
            args.input,
            results,
            gold_records_by_id=gold_records_by_id,
        )
        if args.output_csv.name.startswith("per_sample."):
            review_name = "answer_review." + args.output_csv.name[len("per_sample."):]
        else:
            review_name = "answer_review.csv"
        review_path = args.output_csv.with_name(review_name)
        review_xlsx_path = review_path.with_suffix(".xlsx")
        _preserve_existing_human_eval(answer_review_rows, review_path, review_xlsx_path)
        _write_answer_review_csv(answer_review_rows, review_path)
        print(f"[INFO] Answer review CSV written to: {review_path}")
        if args.write_review_xlsx:
            _write_answer_review_xlsx(answer_review_rows, review_xlsx_path)
            print(f"[INFO] Answer review XLSX written to: {review_xlsx_path}")

    if args.output_summary:
        with args.output_summary.open("w") as f:
            json.dump(summary, f, indent=2, ensure_ascii=False)
        print(f"[INFO] Summary written to: {args.output_summary}")

    if not args.quiet:
        print_summary(summary)


if __name__ == "__main__":
    main()
